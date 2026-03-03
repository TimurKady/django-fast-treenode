# -*- coding: utf-8 -*-
"""
TreeNode Importer Module

This module provides functionality for importing tree-structured data
from various formats, including CSV, JSON, XLSX, YAML, and TSV.

Features:
- Supports field mapping and data type conversion for model compatibility.
- Handles ForeignKey relationships and ManyToMany fields.
- Validates and processes raw data before saving to the database.
- Uses bulk operations for efficient data insertion and updates.
- Supports transactional imports to maintain data integrity.

Version: 3.0.0
Author: Timur Kady
Email: timurkady@yandex.com
"""

# Новый модуль импорта для древовидных структур в Django

import csv
import json
import yaml
import openpyxl
from io import BytesIO, StringIO
from django.db import transaction
from django.core.exceptions import ValidationError, ObjectDoesNotExist

from ..cache import treenode_cache as cache


class TreeNodeImporter:
    """Importer of tree data from various formats."""

    def __init__(self, model, file, file_format):
        """Init."""
        self.model = model
        self.file = file
        self.format = file_format.lower()
        self.rows = []
        self.rows_by_id = {}
        self.result = {"created": 0, "updated": 0, "errors": []}

    def parse(self):
        """Parse a file."""
        if self.format == "xlsx":
            content = self.file.read()  # binary
            self.rows = self._parse_xlsx(content)
        else:
            text = self.file.read()
            if isinstance(text, bytes):
                text = text.decode("utf-8")

            if self.format == "csv":
                self.rows = list(csv.DictReader(StringIO(text)))
            elif self.format == "tsv":
                self.rows = list(csv.DictReader(StringIO(text), delimiter="\t"))
            elif self.format == "json":
                self.rows = json.loads(text)
            elif self.format == "yaml":
                self.rows = yaml.safe_load(text)
            else:
                raise ValueError("Unsupported file format")

        self._build_hierarchy()

    def _parse_xlsx(self, content):
        """Parse the xlsx format."""
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(
            ws.iter_rows(min_row=1, max_row=1))]
        return [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]

    def _build_hierarchy(self):
        """
        Build and check hierarchy.

        Calculates _depth and _path based on parent and priority. Checks that
        each parent exists in either the imported set or the base. The _path
        is built based on priority, as in the main package.
        """
        self.rows_by_id = {str(row.get("id")): row for row in self.rows}

        def build_path_and_depth(row, visited=None):
            if visited is None:
                visited = set()
            row_id = str(row.get("id"))
            if row_id in visited:
                raise ValueError(f"Cycle detected at row {row_id}")
            visited.add(row_id)

            parent_id = str(row.get("parent")) if row.get("parent") else None
            if not parent_id:
                row["_depth"] = 0
                row["_path"] = str(row.get("priority", "0")).zfill(4)
                return row["_path"]

            parent_row = self.rows_by_id.get(parent_id)
            if parent_row:
                parent_path = build_path_and_depth(parent_row, visited)
            else:
                try:
                    self.model.objects.get(pk=parent_id)
                    parent_path = "fromdb"
                except ObjectDoesNotExist:
                    self.result["errors"].append(
                        f"Parent {parent_id} for node {row_id} not found.")
                    parent_path = "invalid"

            row["_depth"] = parent_path.count(
                ".") + 1 if parent_path != "invalid" else 0
            priority = str(row.get("priority", "0")).zfill(4)
            row["_path"] = parent_path + "." + \
                priority if parent_path != "invalid" else priority
            return row["_path"]

        for row in self.rows:
            try:
                build_path_and_depth(row)
            except Exception as e:
                self.result["errors"].append(str(e))

    def import_tree(self):
        """Import tree nodes level by level."""
        with transaction.atomic():
            rows_by_level = {}
            for row in self.rows:
                level = row.get("_depth", 0)
                rows_by_level.setdefault(level, []).append(row)

            for depth in sorted(rows_by_level.keys()):
                to_create = []
                to_update = []
                level_rows = rows_by_level[depth]

                row_ids = [row.get("id") for row in level_rows if row.get("id") is not None]
                existing_by_id = self.model.objects.in_bulk(row_ids)

                for row in rows_by_level[depth]:
                    pk = row.get("id")
                    row_data = dict(row)

                    if "parent" in row_data and (row_data["parent"] == "" or row_data["parent"] is None):
                        row_data["parent"] = None

                    if "parent" in row_data:
                        temp_parent_id = row_data.pop("parent")
                        if temp_parent_id is not None:
                            row_data["parent_id"] = temp_parent_id

                    try:
                        existing_obj = existing_by_id.get(pk)
                        if existing_obj is not None:
                            for field, value in row_data.items():
                                if field in {"id", "_depth", "_path"}:
                                    continue
                                setattr(existing_obj, field, value)

                            existing_obj.full_clean()
                            to_update.append(existing_obj)
                        else:
                            obj = self.model(**row_data)
                            obj.full_clean()
                            to_create.append(obj)
                    except ValidationError as e:
                        self.result["errors"].append(
                            f"Validation error for {pk}: {e}")

                created = self.model.objects.bulk_create(to_create)
                self.result["created"] += len(created)

                for obj in to_update:
                    obj.save()
                self.result["updated"] += len(to_update)

        self.model.tasks.add("update", None)
        cache.invalidate(self.model._meta.label)

        return self.result

    def _get_existing_objects_map(self, rows):
        """Return existing objects keyed by string PK for current level."""
        ids = [row.get("id") for row in rows if row.get("id") not in (None, "")]
        if not ids:
            return {}

        existing_objects = self.model.objects.filter(pk__in=ids)
        return {str(obj.pk): obj for obj in existing_objects}

    def _normalize_parent_reference(self, row, id_map):
        """Normalize parent column to parent_id and resolve temporary IDs."""
        if "parent" in row and (row["parent"] == "" or row["parent"] is None):
            row["parent"] = None

        if "parent" in row:
            temp_parent_id = row.pop("parent")
            if temp_parent_id is not None:
                row["parent_id"] = id_map.get(temp_parent_id, temp_parent_id)

    def _get_update_fields(self, to_update):
        """Return updatable model field names for bulk_update operation."""
        if not to_update:
            return []

        blocked_fields = {"_path", "_depth"}
        model_fields = [
            field.attname for field in self.model._meta.fields
            if not field.primary_key and field.attname not in blocked_fields
        ]
        return model_fields


# The End
