# -*- coding: utf-8 -*-
"""
TreeNode Importer Module

This module provides functionality for importing tree-structured data
from various formats, including CSV, JSON, XLSX, YAML, and TSV.

Features:
- Supports field mapping and data type conversion for model compatibility.
- Handles ForeignKey relationships and ManyToMany fields.
- Validates and processes raw data before saving to the database.
- Uses bulk operations for efficient data insertion and updates.
- Supports transactional imports to maintain data integrity.

Version: 2.0.11
Author: Timur Kady
Email: timurkady@yandex.com
"""


import csv
import json
import yaml
import openpyxl
import math
import uuid
from io import BytesIO, StringIO

import logging

logger = logging.getLogger(__name__)


class TreeNodeImporter:
    """Импортер древовидных данных из различных форматов."""

    def __init__(self, model, file, format, fields=None, mapping=None):
        """
        Init method.

        :param model: Django model where the data will be imported.
        :param file: File object.
        :param format: File format ('csv', 'json', 'xlsx', 'yaml', 'tsv').
        :param fields: List of model fields to import.
        :param mapping: Dictionary for mapping keys from file to model
        field names.
        For example: {"Name": "title", "Description": "desc"}
        """
        self.model = model
        self.format = format
        # Если поля не заданы, используем все поля модели
        self.fields = fields or [field.name for field in model._meta.fields]
        # По умолчанию маппинг идентичен: ключи совпадают с именами полей
        self.mapping = mapping or {field: field for field in self.fields}
        # Считываем содержимое файла один раз, чтобы избежать проблем с курсором
        self.file_content = file.read()

    def get_text_content(self):
        """Return the contents of a file as a string."""
        if isinstance(self.file_content, bytes):
            return self.file_content.decode("utf-8")
        return self.file_content

    def import_data(self):
        """Import data and returns a list of dictionaries."""
        importers = {
            "csv": self.from_csv,
            "json": self.from_json,
            "xlsx": self.from_xlsx,
            "yaml": self.from_yaml,
            "tsv": self.from_tsv,
        }
        if self.format not in importers:
            raise ValueError("Unsupported import format")

        raw_data = importers[self.format]()

        # Processing: field filtering, complex value packing and type casting
        processed = []
        for record in raw_data:
            filtered = self.filter_fields(record)
            filtered = self.process_complex_fields(filtered)
            filtered = self.cast_record_types(filtered)
            processed.append(filtered)

        return processed

    def get_tn_orders(self, rows):
        """Calculate the materialized path without including None parents."""
        # Build a mapping from id to record for quick lookup.
        row_dict = {row["id"]: row for row in rows}

        def get_ancestor_path(row):
            parent_field = 'tn_parent' if 'tn_parent' in row else 'tn_parent_id'
            return get_ancestor_path(row_dict[row[parent_field]]) + [row["id"]] if row[parent_field] else [row["id"]]

        return [
            {"id": row["id"], "path": get_ancestor_path(row)}
            for row in rows
        ]

    def filter_fields(self, record):
        """
        Filter the record according to the mapping.

        Only the necessary keys remain, while the names are renamed.
        """
        new_record = {}
        for file_key, model_field in self.mapping.items():
            new_record[model_field] = record.get(file_key)
        return new_record

    def process_complex_fields(self, record):
        """
        Pack it into a JSON string.

        If the field value is a dictionary or list.
        """
        for key, value in record.items():
            if isinstance(value, uuid.UUID):
                record[key] = str(value)
            if isinstance(value, (list, dict)):
                try:
                    record[key] = json.dumps(value, ensure_ascii=False)
                except Exception as e:
                    logger.warning("Error serializing field %s: %s", key, e)
                    record[key] = None
        return record

    def cast_record_types(self, record):
        """
        Cast the values ​​of the record fields to the types defined in the model.

        For each field, its to_python() method is called. If the value is nan,
        it is replaced with None.
        For ForeignKey fields (many-to-one), the value is written to
        the <field>_id attribute, and the original key is removed.
        """
        for field in self.model._meta.fields:
            field_name = field.name
            if field.is_relation and field.many_to_one:
                if field_name in record:
                    value = record[field_name]
                    if isinstance(value, float) and math.isnan(value):
                        value = None
                    try:
                        converted = None if value is None else int(value)
                        # Записываем в атрибут, например, tn_parent_id
                        record[field.attname] = converted
                    except Exception as e:
                        logger.warning(
                            "Error converting FK field %s with value %r: %s",
                            field_name,
                            value,
                            e
                        )
                        record[field.attname] = None
                    # Удаляем оригинальное значение, чтобы Django не пыталась
                    # его обработать
                    del record[field_name]
            else:
                if field_name in record:
                    value = record[field_name]
                    if isinstance(value, float) and math.isnan(value):
                        record[field_name] = None
                    else:
                        try:
                            record[field_name] = field.to_python(value)
                        except Exception as e:
                            logger.warning(
                                "Error converting field %s with value %r: %s",
                                field_name,
                                value,
                                e
                            )
                            record[field_name] = None
        return record

    # ------------------------------------------------------------------------

    def finalize(self, raw_data):
        """
        Finalize import.

        Processes raw_data, creating and updating objects by levels
        (from roots to leaves) using the materialized path to calculate
        the level.

        Algorithm:
        1. Build a raw_by_id dictionary for quick access to records by id.
        2. For each record, calculate the materialized path:
        - If tn_parent is specified and exists in raw_data, recursively get
          the parent's path and add its id.
        - If tn_parent is missing from raw_data, check if the parent is in
          the database.
        If not, generate an error.
        3. Record level = length of its materialized path.
        4. Split records into those that need to be created (if the object
           with the given id is not yet in the database), and those that need
           to be updated.
        5. To create, process groups by levels (sort by increasing level):
        - Validate each record, if there are no errors, add the instance to
          the list.
        - After each level, we perform bulk_create.
        6. For updates, we collect instances, fixing fields (without id)
           and perform bulk_update.

        Returns a dictionary:
          {
             "create": [созданные объекты],
             "update": [обновлённые объекты],
             "errors": [список ошибок]
          }
        """
        result = {
            "create": [],
            "update": [],
            "errors": []
        }

        # 1. Calculate the materialized path and level for each entry.
        paths = self.get_tn_orders(raw_data)
        # key: record id, value: уровень (int)
        levels_by_record = {rec["id"]: len(rec["path"])-1 for rec in paths}

        # 2. Разбиваем записи по уровням
        levels = {}
        for record in raw_data:
            level = levels_by_record.get(record["id"], 0)
            if level not in levels:
                levels[level] = []
            levels[level].append(record)

        records_by_level = [
            sorted(
                levels[key],
                key=lambda x: (x.get(
                    "tn_parent",
                    x.get("tn_parent_id", 0)) or -1)
            )
            for key in sorted(levels.keys())
        ]

        # 4. We split the records into those to create and those to update.
        # The list of records to update
        to_update = []

        for level in range(len(records_by_level)):
            instances_to_create = []
            for record in records_by_level[level]:
                rec_id = record["id"]
                if self.model.objects.filter(pk=rec_id).exists():
                    to_update.append(record)
                else:
                    instance = self.model(**record)
                    try:
                        instance.full_clean()
                        instances_to_create.append(instance)
                    except Exception as e:
                        result["errors"].append(f"Validation error for record \
{record['id']} on level {level}: {e}")
            try:
                created = self.model.objects.bulk_create(instances_to_create)
                result["create"].extend(created)
            except Exception as e:
                result["errors"].append(f"Create error on level {level}: {e}")

        # 6. Processing updates: collecting instances and a list of fields
        # for bulk_update
        updated_instances = []
        update_fields_set = set()
        for record in to_update:
            rec_id = record["id"]
            try:
                instance = self.model.objects.get(pk=rec_id)
                for field, value in record.items():
                    if field != "id":
                        setattr(instance, field, value)
                        update_fields_set.add(field)
                instance.full_clean()
                updated_instances.append(instance)
            except Exception as e:
                result["errors"].append(
                    f"Validation error updating record {rec_id}: {e}")
        update_fields = list(update_fields_set)
        if updated_instances:
            try:
                self.model.objects.bulk_update(updated_instances, update_fields)
                result["update"].extend(updated_instances)
            except Exception as e:
                result["errors"].append(f"Bulk update error: {e}")

        return result

    # ------------------------------------------------------------------------

    def from_csv(self):
        """Import from CSV."""
        text = self.get_text_content()
        return list(csv.DictReader(StringIO(text)))

    def from_json(self):
        """Import from JSON."""
        return json.loads(self.get_text_content())

    def from_xlsx(self):
        """Import from XLSX (Excel)."""
        file_stream = BytesIO(self.file_content)
        rows = []
        wb = openpyxl.load_workbook(file_stream, read_only=True)
        ws = wb.active
        headers = [
            cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows

    def from_yaml(self):
        """Import from YAML."""
        return yaml.safe_load(self.get_text_content())

    def from_tsv(self):
        """Import from TSV."""
        text = self.get_text_content()
        return list(csv.DictReader(StringIO(text), delimiter="\t"))

# The End
